from openpyxl import load_workbook
import openpyxl;
#import write;

#wb=load_workbook(filename='D:/EMOS工单.xlsx')
#print(wb.get_sheet_names())
#读取excel文件
def readwb(wbname,sheetname):
    wb=openpyxl.load_workbook(filename=wbname,read_only=True)
    if (sheetname==""):
        ws=wb.active
    else:
        ws=wb[sheetname]
    data=[]
    for row in ws.rows:
        list=[]
        for cell in row:
            aa=str(cell.value)
            if (aa=="1"):
                aa="10"
            list.append(aa)
        data.append(list)

    print (wbname +"-"+sheetname+"- 已成功读取")
    return data
readwb('DDS工单.xlsx','Sheet1')
