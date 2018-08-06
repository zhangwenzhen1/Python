from openpyxl import load_workbook  
#from openpyxl.styles import Color, Font, Alignment  
#from openpyxl.styles.fills import FILL_SOLID  
#from openpyxl.styles.colors import BLUE, RED, GREEN, YELLOW  


#wb=load_workbook(filename='d:/EMOS工单-公式 -中兴&爱立信-小时粒度版.xlsx')
#print(wb.get_sheet_names())

wb = load_workbook(filename='D:/DDS工单.xlsx')
# 获得所有sheet的名称
print(wb.get_sheet_names())
# 根据sheet名字获得sheet
a_sheet = wb.get_sheet_by_name('Sheet5')
# 获得sheet名
print(a_sheet.title)
# 获得当前正在显示的sheet, 也可以用wb.get_active_sheet()
a_sheet = wb.active
#sheet=wb.get_active_sheet('Sheet4')
# 获取某个单元格的值，观察excel发现也是先字母再数字的顺序，即先列再行
b4 = a_sheet['B4']
row1 = [1, 2, 3, 4, 5]
a_sheet.append(row1)
# 分别返回
print(f'({b4.column}, {b4.row}) is {b4.value}') # 返回的数字就是int型
# 除了用下标的方式获得，还可以用cell函数, 换成数字，这个表示B4
b4_too = a_sheet.cell(row=4, column=2)
print(b4_too.value)
print(a_sheet.max_row)
print(a_sheet.max_column)
b_sheet = wb.get_sheet_by_name('Sheet4')
#b_sheet = wb.get_active_sheet('Sheet4')
#b_sheet = wb.active
row = [1, 2, 3, 4, 5]
b_sheet.append(row)
b_sheet['A6'] = 'good'
print(b_sheet.max_row)
wb.save('D:/DDS工单.xlsx')
#sheet.rows
#row_length=len(sheet.rows)
"""
b_sheet = wb.get_sheet_by_name('Sheet4')
row = [1 ,2, 3, 4, 5]
b_sheet.append(row)
b_sheet['A2'] = 'good'
print(b_sheet.max_row)
wb.save(r'D:/DDS工单.xlsx')
"""

"""
txt1=u'该小区存在'
ws['A2'] = txt1
a=str(ws['A1'].value)
b=ws['A2'].value
ws['A3']=a+b
ws1=wb.get_sheet_by_name('第一')
ws1['A2']=2333
ws1['A3']=66
ws1['A4']=ws1['A2'].value+ws1['A3'].value
wb.save('d:/DDS工单.xlsx')
"""

