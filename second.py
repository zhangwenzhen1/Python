from openpyxl import load_workbook

wb = load_workbook(filename='D:/DDS工单.xlsx')
# 获得所有sheet的名称
print(wb.get_sheet_names())
# 根据sheet名字获得sheet
a_sheet = wb.get_sheet_by_name('Sheet5')
# 获得sheet名
print(a_sheet.title)
# 获得当前正在显示的sheet, 也可以用wb.get_active_sheet()
a_sheet = wb.active
#sheet=wb.get_active_sheet('Sheet4')
# 获取某个单元格的值，观察excel发现也是先字母再数字的顺序，即先列再行
b4 = a_sheet['B4']
row1 = [1, 2, 3, 4, 5]
a_sheet.append(row1)
# 分别返回
print(f'({b4.column}, {b4.row}) is {b4.value}') # 返回的数字就是int型
# 除了用下标的方式获得，还可以用cell函数, 换成数字，这个表示B4
b4_too = a_sheet.cell(row=4, column=2)
print(b4_too.value)
print(a_sheet.max_row)
print(a_sheet.max_column)
wb.save('D:/DDS工单.xlsx')